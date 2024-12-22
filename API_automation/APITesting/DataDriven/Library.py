import json
import jsonpath
import requests
import openpyxl

class Common:
    def __init__(self,Filepath,SheetName):
        global vk
        global sh
        vk = openpyxl.load_workbook(Filepath)
        sh = vk[SheetName]
    
    def fetch_row_count(self):
        #Reading data from xls file
        rows = sh.max_row
        print("no of rows:", rows)
        return rows
    def fetch_column_count(self, ):
        col = sh.max_column
        return col
    def fetch_key_names(self):
        c = sh.max_column
        li=[]
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            li.insert(i-1, cell.value)
        return li
    
    def update_request_with_data(self,rowNumber,jsonrequest,keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonrequest[i-1] = cell.value
            
        return jsonrequest