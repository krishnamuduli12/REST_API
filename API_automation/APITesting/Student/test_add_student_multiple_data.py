import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_student():
    url = 'https://thetestingworldapi.com/api/studentsDetails'
    f = open("C:\\workspace\\Learning\\Python\\API_automation\\APITesting\\Student\\AddNewStudent.json","r")
    json_request = json.loads(f.read())
    
    #Reading data from xls file
    vk = openpyxl.load_workbook("C:\\workspace\\Learning\\Python\\API_automation\\APITesting\\Student\\test_data.xlsx")
    sh = vk['Sheet1']
    rows = sh.max_row
    print("no of rows:", rows)
    
    for i in range(2, rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['mid_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value
        response = requests.post(url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201 
    
       
    
    
    
    